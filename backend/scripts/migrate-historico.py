"""
Script de migração dos dados históricos do Excel/CSV para o PostgreSQL.
Lê os arquivos em docs/ e insere clientes, motoristas, cargas e cotações.
"""

import os
import re
import csv
import sys
from datetime import date, datetime
from pathlib import Path
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ─── Configuração ────────────────────────────────────────────────────────────

BACKEND_DIR = Path(__file__).parent.parent
DOCS_DIR    = BACKEND_DIR.parent / "docs"
EXCEL_FILE  = DOCS_DIR / "DESEMPENHO PESSOAL - THIAGO.xlsx"
CSV_FILE    = DOCS_DIR / "COTAÇOES THIAGO - COTAÇÃO.csv"

# Lê DATABASE_URL do .env
env_path = BACKEND_DIR / ".env"
DB_URL = None
if env_path.exists():
    for line in env_path.read_text(encoding="utf-8").splitlines():
        if line.startswith("DATABASE_URL="):
            DB_URL = line.split("=", 1)[1].strip()
            break

if not DB_URL:
    print("ERRO: DATABASE_URL não encontrado no .env")
    sys.exit(1)

# ─── Helpers ─────────────────────────────────────────────────────────────────

def normalizar_nome(nome: str) -> str:
    """Remove espaços extras e converte para maiúsculas."""
    return " ".join(nome.strip().upper().split())

def parse_decimal(val) -> float | None:
    """Converte valor de célula Excel ou string para float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = s.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None

def parse_decimal_req(val, default=0.0) -> float:
    """parse_decimal mas com fallback para valor padrão."""
    r = parse_decimal(val)
    return r if r is not None else default

def parse_data(val) -> date | None:
    """Converte valor de data do Excel para date."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def parse_pct_csv(val) -> float | None:
    """Converte '12,00%' → 12.0 (formato human-readable para cotações)."""
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip().replace("%", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None

def parse_moeda_csv(val) -> float | None:
    """Converte 'R$ 900,00' → 900.0"""
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip()
    s = re.sub(r"[R$\s]", "", s)
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

# ─── Mapeamento de colunas do Excel ──────────────────────────────────────────

def build_col_map(headers: list[str]) -> dict:
    """
    Constrói mapa nome→índice a partir da linha de cabeçalho.
    Lida com duplicatas (ex: dois 'SEGURO', dois 'ICMS', dois 'IMPOSTO').
    """
    clean = [str(h).strip().upper() if h else "" for h in headers]
    seen: dict[str, int] = {}
    result: dict[str, int] = {}

    for i, h in enumerate(clean):
        if not h:
            continue
        count = seen.get(h, 0)
        if count == 0:
            result[h] = i
        else:
            result[f"{h}_{count+1}"] = i
        seen[h] = count + 1

    return result

def get_col(cm: dict, *names: str):
    """Retorna o índice do primeiro nome encontrado, ou None."""
    for n in names:
        if n in cm:
            return cm[n]
    return None

# ─── Parser de aba do Excel ───────────────────────────────────────────────────

ABAS_MESES = [
    # (nome_aba, ano_esperado_ou_None)
    ("JANEIRO",          2025),
    ("FEVEREIRO",        2025),
    ("MARÇO",            2025),
    ("ABRIL",            2025),
    ("MAIO",             2025),
    ("JUNHO",            2025),
    ("JULHO",            2025),
    ("AGOSTO",           2025),
    ("SETEMBRO",         2025),
    ("OUTUBRO",          2025),
    ("NOVEMBRO",         2025),
    ("DEZEMBRO",         2025),
    ("JANEIRO - 2026",   2026),
    ("FEVEREIRO - 2026", 2026),
    ("MARÇO - 2026",     None),   # Contém tanto 2025 quanto 2026 — aceitar todos
    ("ABRIL - 2026",     2026),
    ("MAIO - 2026",      2026),
    ("JUNHO - 2026",     2026),
]

def parse_aba_cargas(ws, nome_aba: str) -> list[dict]:
    """Lê uma aba de mês e retorna lista de dicts com dados de cada carga."""
    rows_iter = ws.iter_rows(values_only=True)
    header_row = None

    for i, row in enumerate(rows_iter, 1):
        if i == 3:
            header_row = list(row)
        if i >= 4 and header_row is not None:
            break

    if header_row is None:
        return []

    cm = build_col_map(header_row)

    # Índices com fallback entre nomes alternativos
    idx_data      = get_col(cm, "DATA")
    idx_cte       = get_col(cm, "CTE")
    idx_origem    = get_col(cm, "ORIGEM")
    idx_destino   = get_col(cm, "DESTINO")
    idx_cliente   = get_col(cm, "CLIENTE")
    idx_pagamento = get_col(cm, "PAGAMENTO")
    idx_canhoto   = get_col(cm, "CANHOTO")
    idx_motorista = get_col(cm, "MOTORISTA")
    idx_tel       = get_col(cm, "TELEFONE")
    idx_val_mot   = get_col(cm, "VALOR MOTORISTA")
    idx_val_emp   = get_col(cm, "VALOR EMPRESA")
    idx_nf        = get_col(cm, "NF")

    # SEGURO: % e R$ (GRIS em abas antigas, SEGURO em novas)
    idx_seg_p  = get_col(cm, "GRIS", "SEGURO")
    idx_seg_v  = get_col(cm, "SEGURO_2", "SEGURO")  # segunda ocorrência
    if idx_seg_p == idx_seg_v:
        idx_seg_v = None  # abas antigas só têm um (R$ inline)

    # ICMS: % e R$ (segunda ocorrência = R$)
    idx_icms_p = get_col(cm, "ICMS")
    idx_icms_v = cm.get("ICMS_2")  # segunda ocorrência, pode não existir

    # C.O / CO: % e R$
    idx_co_p   = get_col(cm, "C.O")
    idx_co_v   = get_col(cm, "CO")
    if idx_co_v == idx_co_p:
        idx_co_v = None

    # IMPOSTO: % e R$
    idx_imp_p  = get_col(cm, "IMPOSTO")
    idx_imp_v  = cm.get("IMPOSTO_2")

    # BOLETO (pode ser ANTECIPAÇÃO em abas antigas)
    idx_boleto_p = get_col(cm, "BOLETO", "ANTECIPAÇÃO", "ANTECIPACAO")

    idx_dia_pag     = get_col(cm, "DIA PAG")
    idx_total_taxas = get_col(cm, "TOTAL TAXAS")
    idx_val_imp_tot = get_col(cm, "VALOR IMPOSTO")
    idx_pct_com     = get_col(cm, "%")
    idx_com_v       = get_col(cm, "COMISSÕES", "COMISSOES")
    idx_bol_v       = get_col(cm, "BOLETOS")
    idx_lucro       = get_col(cm, "LUCRO")
    idx_rentab      = get_col(cm, "RENTABIL")
    idx_mova        = get_col(cm, "MOVA")
    idx_adq         = get_col(cm, "VALOR")  # última coluna = ADQ TAB R$

    cargas = []

    # Continuar lendo as linhas de dados (já passamos da linha 3)
    # Re-iterar a partir da linha 4 exige re-abrir; vamos usar rows completas
    return cargas  # placeholder — lemos abaixo via nova iteração

def ler_aba_cargas(ws, nome_aba: str) -> list[dict]:
    """Lê aba completa e retorna lista de dicts de cargas."""
    all_rows = list(ws.iter_rows(values_only=True))

    # Linha 3 (índice 2) é o cabeçalho real
    if len(all_rows) < 3:
        return []

    header_row = list(all_rows[2])
    cm = build_col_map(header_row)

    idx_data      = get_col(cm, "DATA")
    idx_cte       = get_col(cm, "CTE")
    idx_origem    = get_col(cm, "ORIGEM")
    idx_destino   = get_col(cm, "DESTINO")
    idx_cliente   = get_col(cm, "CLIENTE")
    idx_pagamento = get_col(cm, "PAGAMENTO")
    idx_canhoto   = get_col(cm, "CANHOTO")
    idx_motorista = get_col(cm, "MOTORISTA")
    idx_val_mot   = get_col(cm, "VALOR MOTORISTA")
    idx_val_emp   = get_col(cm, "VALOR EMPRESA")
    idx_nf        = get_col(cm, "NF")

    # SEGURO: dois padrões
    # Abas antigas: GRIS (%) + SEGURO (R$) — 1 ocorrência de cada
    # Abas novas: SEGURO (%) + SEGURO_2 (R$)
    idx_seg_p = cm.get("GRIS", cm.get("SEGURO"))
    idx_seg_v = cm.get("SEGURO_2") if "GRIS" not in cm else cm.get("SEGURO")

    idx_icms_p  = cm.get("ICMS")
    idx_icms_v  = cm.get("ICMS_2")

    idx_co_p    = cm.get("C.O")
    idx_co_v    = cm.get("CO")
    if idx_co_v is not None and idx_co_v == idx_co_p:
        idx_co_v = None

    idx_imp_p   = cm.get("IMPOSTO")
    idx_imp_v   = cm.get("IMPOSTO_2")

    idx_boleto_p = get_col(cm, "BOLETO", "ANTECIPAÇÃO", "ANTECIPACAO")
    idx_dia_pag     = cm.get("DIA PAG")
    idx_total_taxas = cm.get("TOTAL TAXAS")
    idx_val_imp_tot = cm.get("VALOR IMPOSTO")
    idx_pct_com     = cm.get("%")
    idx_com_v       = get_col(cm, "COMISSÕES", "COMISSOES")
    idx_bol_v       = cm.get("BOLETOS")
    idx_lucro       = cm.get("LUCRO")
    idx_rentab      = cm.get("RENTABIL")
    idx_mova        = cm.get("MOVA")
    idx_adq_v       = cm.get("VALOR")

    def safe(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    cargas = []
    for row in all_rows[3:]:  # a partir da linha 4
        if not row or len(row) == 0:
            continue
        data_val = parse_data(safe(row, idx_data))
        if data_val is None:
            continue  # linha vazia ou inválida

        cliente_raw = safe(row, idx_cliente)
        if not cliente_raw or not str(cliente_raw).strip():
            continue

        val_mot = parse_decimal(safe(row, idx_val_mot))
        val_emp = parse_decimal(safe(row, idx_val_emp))
        if val_mot is None or val_emp is None:
            continue  # linha sem valores — provavelmente subtotal

        motorista_raw = safe(row, idx_motorista)
        motorista_nome = normalizar_nome(str(motorista_raw)) if motorista_raw else None

        seg_p = parse_decimal_req(safe(row, idx_seg_p))
        seg_v_raw = parse_decimal(safe(row, idx_seg_v))
        val_nf = parse_decimal_req(safe(row, idx_nf))

        # Se seg_v não existe (abas antigas), calcula
        seg_v = seg_v_raw if seg_v_raw is not None else (seg_p * val_nf if val_nf else 0.0)

        icms_p = parse_decimal_req(safe(row, idx_icms_p))
        co_p   = parse_decimal_req(safe(row, idx_co_p))
        imp_p  = parse_decimal_req(safe(row, idx_imp_p))

        co_v   = parse_decimal_req(safe(row, idx_co_v))  if idx_co_v  is not None else co_p  * val_emp
        imp_v  = parse_decimal_req(safe(row, idx_imp_v)) if idx_imp_v is not None else imp_p * val_emp

        boleto_p     = parse_decimal_req(safe(row, idx_boleto_p))
        dias_pag     = int(parse_decimal_req(safe(row, idx_dia_pag), 15))
        total_taxas  = parse_decimal_req(safe(row, idx_total_taxas))
        val_imp_tot  = parse_decimal_req(safe(row, idx_val_imp_tot))
        pct_com      = parse_decimal_req(safe(row, idx_pct_com))
        com_v        = parse_decimal_req(safe(row, idx_com_v))
        bol_v        = parse_decimal_req(safe(row, idx_bol_v), 17.5)
        lucro        = parse_decimal_req(safe(row, idx_lucro))
        rentab_raw   = parse_decimal_req(safe(row, idx_rentab))
        rentab       = rentab_raw if abs(rentab_raw) < 99 else 0.0  # clampa erro de dados
        mova_v       = parse_decimal_req(safe(row, idx_mova))
        adq_v        = parse_decimal_req(safe(row, idx_adq_v))

        pagamento    = str(safe(row, idx_pagamento) or "").strip().upper()
        status       = "entregue" if "FINALIZADO" in pagamento else "em_andamento"

        canhoto_raw  = str(safe(row, idx_canhoto) or "").strip().upper()
        canhoto      = canhoto_raw in ("SIM", "POSTADO", "1", "TRUE")

        cte_raw = safe(row, idx_cte)
        cte = str(cte_raw).strip() if cte_raw is not None else None
        if cte and (cte == "None" or cte == "nan"):
            cte = None

        origem  = str(safe(row, idx_origem) or "").strip().upper() or None
        destino = str(safe(row, idx_destino) or "").strip().upper() or None

        cargas.append({
            "aba":               nome_aba,
            "data":              data_val,
            "cte":               cte,
            "origem":            origem,
            "destino":           destino,
            "cliente":           normalizar_nome(str(cliente_raw)),
            "motorista":         motorista_nome,
            "valor_motorista":   val_mot,
            "valor_empresa":     val_emp,
            "valor_nf":          val_nf,
            "seguro_percent":    seg_p,
            "seguro_valor":      seg_v,
            "icms_percent":      icms_p,
            "co_percent":        co_p,
            "co_valor":          co_v,
            "imposto_percent":   imp_p,
            "imposto_valor":     imp_v,
            "boleto_percent":    boleto_p,
            "dias_pagamento":    dias_pag,
            "total_taxas_percent": total_taxas,
            "valor_imposto_total": val_imp_tot,
            "percent_comissao":  pct_com,
            "comissao_valor":    com_v,
            "boletos_valor":     bol_v,
            "lucro":             lucro,
            "percent_rentabilidade": rentab,
            "mova_valor":        mova_v,
            "adq_tab_valor":     adq_v,
            "status":            status,
            "canhoto_enviado":   canhoto,
        })

    return cargas


def ler_cotacoes_csv(path: Path) -> list[dict]:
    """Lê o CSV de cotações e retorna lista de dicts."""
    cotacoes = []
    with open(path, encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader)
        # Normaliza header
        hdr = [h.strip().upper() for h in header]

        # Encontrar índices
        def ci(name):
            try:
                return hdr.index(name)
            except ValueError:
                return None

        idx_data    = ci("DATA")
        idx_cli     = ci("CLIENTE")
        idx_orig    = ci("ORIGEM")
        idx_dest    = ci("DESTINO")
        idx_km      = ci("KM")
        idx_veic    = ci("VEÍCULO") or ci("VEICULO")
        idx_mot     = ci("R$ MOTORISTA")
        idx_emp     = ci("R$ EMPRESA")
        idx_nf      = ci("VALOR DE NF")
        idx_dias    = ci("DIAS PAG")
        idx_icms    = ci("ICMS")
        idx_co      = ci("CO")
        idx_com     = ci("COMISSÃO") or ci("COMISSAO")
        idx_rent    = ci("RENTABILIDADE")
        idx_vcom    = ci("V. COM")
        idx_sit     = ci("SITUAÇÃO") or ci("SITUACAO")

        for row in reader:
            if not row or len(row) < 2:
                continue

            def g(idx):
                if idx is None or idx >= len(row):
                    return None
                v = row[idx].strip()
                return v if v else None

            data_raw = g(idx_data)
            cliente_raw = g(idx_cli)
            if not data_raw or not cliente_raw:
                continue

            # Parse data dd/mm/yyyy
            data_val = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    data_val = datetime.strptime(data_raw, fmt).date()
                    break
                except ValueError:
                    continue
            if data_val is None:
                continue

            val_mot = parse_moeda_csv(g(idx_mot))
            val_emp = parse_moeda_csv(g(idx_emp))
            if val_emp is None or val_emp == 0:
                continue

            val_nf  = parse_moeda_csv(g(idx_nf))
            km_raw  = g(idx_km)
            km      = None
            if km_raw:
                km = parse_decimal(km_raw.replace(".", "").replace(",", "."))

            icms_p  = parse_pct_csv(g(idx_icms))   # já em % (ex: 12.0)
            co_p    = parse_pct_csv(g(idx_co))
            com_p   = parse_pct_csv(g(idx_com))
            rent    = parse_pct_csv(g(idx_rent))    # ex: 15,04 → 15.04
            vcom    = parse_moeda_csv(g(idx_vcom))

            sit_raw = str(g(idx_sit) or "").strip().upper()
            situacao = "batida" if "BATIDA" in sit_raw else "pendente"

            dias_raw = g(idx_dias)
            dias = int(parse_decimal_req(dias_raw, 15)) if dias_raw else 15

            # Calcular percentRentabilidade como decimal (divide por 100)
            pct_rent = (rent / 100.0) if rent is not None else 0.0

            cotacoes.append({
                "data":               data_val.isoformat(),
                "cliente":            normalizar_nome(str(cliente_raw)),
                "origem":             str(g(idx_orig) or "").strip().upper() or None,
                "destino":            str(g(idx_dest) or "").strip().upper() or None,
                "km":                 km,
                "tipo_veiculo":       str(g(idx_veic) or "").strip() or None,
                "valor_motorista":    val_mot,
                "valor_empresa":      val_emp,
                "valor_nf":           val_nf,
                "icms_percent":       icms_p or 4.0,      # human-readable para cotações
                "co_percent":         co_p or 3.5,
                "imposto_percent":    6.5,
                "seguro_percent":     0.03,
                "dias_pagamento":     dias,
                "percent_comissao":   com_p or 10.0,
                "comissao_valor":     vcom or 0.0,
                "lucro":              0.0,
                "percent_rentabilidade": pct_rent,
                "situacao":           situacao,
            })

    return cotacoes


# ─── Migração principal ───────────────────────────────────────────────────────

def main():
    print("=== Migracao de Dados Historicos ===\n")

    # 1. Ler dados do Excel
    print("[1/6] Lendo planilha Excel...")
    wb = openpyxl.load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)

    todas_cargas: list[dict] = []
    for nome_aba, _ in ABAS_MESES:
        # Tentar variantes de encoding (MARÇO pode aparecer como MAR\xc7O)
        ws = None
        for candidate in [nome_aba, nome_aba.replace("Ç", "\xc7").replace("Ã", "\xc3")]:
            if candidate in wb.sheetnames:
                ws = wb[candidate]
                break
        if ws is None:
            # Busca por substring normalizada
            for sname in wb.sheetnames:
                sn = sname.encode("latin-1", errors="replace").decode("utf-8", errors="replace")
                if nome_aba[:5].upper() in sn.upper():
                    ws = wb[sname]
                    break
        if ws is None:
            print(f"  AVISO: Aba '{nome_aba}' nao encontrada -- ignorando")
            continue

        cargas_aba = ler_aba_cargas(ws, nome_aba)
        print(f"  {nome_aba:25s}: {len(cargas_aba):4d} cargas")
        todas_cargas.extend(cargas_aba)

    print(f"\n  Total bruto: {len(todas_cargas)} cargas")

    # 2. Ler cotacoes do CSV
    print("\n[2/6] Lendo CSV de cotacoes...")
    cotacoes = ler_cotacoes_csv(CSV_FILE)
    print(f"  Total: {len(cotacoes)} cotacoes")

    # 3. Deduplicate cargas por CTE (quando CTE existe)
    print("\n[3/6] Deduplicando cargas por CTE...")
    seen_ctes: set[str] = set()
    cargas_dedup: list[dict] = []
    sem_cte = 0
    for c in todas_cargas:
        cte = c.get("cte")
        if cte:
            cte_key = cte.strip()
            if cte_key in seen_ctes:
                continue
            seen_ctes.add(cte_key)
        else:
            sem_cte += 1
        cargas_dedup.append(c)
    print(f"  Apos dedup: {len(cargas_dedup)} cargas ({sem_cte} sem CTE)")

    # 4. Extrair clientes unicos
    clientes_nomes: set[str] = set()
    for c in cargas_dedup:
        clientes_nomes.add(c["cliente"])
    for ct in cotacoes:
        clientes_nomes.add(ct["cliente"])
    clientes_lista = sorted(clientes_nomes)
    print(f"\n  Clientes unicos: {len(clientes_lista)}")

    # 5. Extrair motoristas unicos (normalizado)
    motoristas_map: dict[str, str | None] = {}
    for c in cargas_dedup:
        mot = c.get("motorista")
        if mot and mot not in motoristas_map:
            motoristas_map[mot] = None
    motoristas_lista = sorted(motoristas_map.keys())
    print(f"  Motoristas unicos: {len(motoristas_lista)}")
    print(f"  (+ 'MOTORISTA NAO IDENTIFICADO' para cargas sem motorista)")

    # 6. Conectar ao banco
    print(f"\n[4/6] Conectando ao PostgreSQL...")
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur = conn.cursor()

    try:
        # 7. Garantir que as tabelas existem (schema já foi pushed via drizzle)
        # Verificar tabelas
        cur.execute("SELECT tablename FROM pg_tables WHERE schemaname='public'")
        tabelas_existentes = {r[0] for r in cur.fetchall()}
        tabelas_req = {"clientes", "motoristas", "cargas", "cotacoes", "indicadores_semanais"}
        faltando = tabelas_req - tabelas_existentes
        if faltando:
            print(f"\nERRO: Tabelas faltando: {faltando}")
            print("   Execute 'pnpm db:push' no backend antes de rodar esta migracao.")
            return

        # 8. Limpar dados existentes (re-rodar idempotente)
        print("\n[5/6] Limpando dados existentes...")
        cur.execute("DELETE FROM cargas")
        cur.execute("DELETE FROM cotacoes")
        cur.execute("DELETE FROM indicadores_semanais")
        cur.execute("DELETE FROM metas_mensais")
        cur.execute("DELETE FROM motoristas")
        cur.execute("DELETE FROM clientes")
        conn.commit()
        print("   OK.")

        # 9. Inserir clientes
        print("\n[6/6] Inserindo dados...")
        print("  Inserindo clientes...")
        cur.execute("SELECT setval(pg_get_serial_sequence('clientes','id'), 1, false)")
        cliente_rows = [(nome,) for nome in clientes_lista]
        execute_values(cur,
            "INSERT INTO clientes (nome, created_at) VALUES %s RETURNING id, nome",
            [(nome, "now()") for nome in clientes_lista]
        )
        rows = cur.fetchall()
        cliente_id: dict[str, int] = {r[1]: r[0] for r in rows}
        print(f"   {len(cliente_id)} clientes inseridos.")

        # 10. Inserir motoristas (incluindo placeholder)
        print("  Inserindo motoristas...")
        placeholder_nome = "MOTORISTA NÃO IDENTIFICADO"
        motoristas_para_inserir = [placeholder_nome] + motoristas_lista
        mot_rows = []
        for nome in motoristas_para_inserir:
            mot_rows.append((nome, nome))  # (nome, nome_normalizado)

        execute_values(cur,
            """INSERT INTO motoristas (nome, nome_normalizado, created_at)
               VALUES %s
               ON CONFLICT (nome_normalizado) DO NOTHING
               RETURNING id, nome_normalizado""",
            [(nome, nome_norm, "now()") for nome, nome_norm in mot_rows]
        )
        # Re-buscar IDs (ON CONFLICT pode ter ignorado alguns)
        cur.execute("SELECT id, nome_normalizado FROM motoristas")
        motorista_id: dict[str, int] = {r[1]: r[0] for r in cur.fetchall()}
        placeholder_id = motorista_id[placeholder_nome]
        print(f"   {len(motorista_id)} motoristas inseridos.")

        # 11. Inserir cargas
        print("  Inserindo cargas...")
        cargas_ok = 0
        cargas_skip = 0
        cargas_batch = []

        for c in cargas_dedup:
            cli_id = cliente_id.get(c["cliente"])
            if cli_id is None:
                cargas_skip += 1
                continue

            mot_nome = c.get("motorista")
            mot_id   = motorista_id.get(mot_nome) if mot_nome else None
            if mot_id is None:
                mot_id = placeholder_id

            cargas_batch.append((
                c["data"].isoformat(),
                c.get("cte"),
                c.get("origem"),
                c.get("destino"),
                cli_id,
                mot_id,
                c["valor_motorista"],
                c["valor_empresa"],
                c["valor_nf"],
                c["seguro_percent"],
                c["icms_percent"],
                c["co_percent"],
                c["imposto_percent"],
                c["dias_pagamento"],
                c["percent_comissao"],
                c["seguro_valor"],
                c["co_valor"],
                c["imposto_valor"],
                c["boleto_percent"],
                c["mova_valor"],
                c["adq_tab_valor"],
                c["total_taxas_percent"],
                c["valor_imposto_total"],
                c["comissao_valor"],
                c["boletos_valor"],
                c["lucro"],
                c["percent_rentabilidade"],
                c["status"],
                c["canhoto_enviado"],
            ))

        if cargas_batch:
            execute_values(cur, """
                INSERT INTO cargas (
                    data, cte, origem, destino,
                    cliente_id, motorista_id,
                    valor_motorista, valor_empresa, valor_nf,
                    seguro_percent, icms_percent, co_percent, imposto_percent,
                    dias_pagamento, percent_comissao,
                    seguro_valor, co_valor, imposto_valor,
                    boleto_percent, mova_valor, adq_tab_valor,
                    total_taxas_percent, valor_imposto_total,
                    comissao_valor, boletos_valor, lucro,
                    percent_rentabilidade, status, canhoto_enviado
                ) VALUES %s
            """, cargas_batch)
            cargas_ok = len(cargas_batch)

        print(f"   {cargas_ok} cargas inseridas, {cargas_skip} ignoradas.")

        # 12. Inserir cotacoes
        print("  Inserindo cotacoes...")
        cot_ok = 0
        cot_skip = 0
        cot_batch = []

        for ct in cotacoes:
            cli_id = cliente_id.get(ct["cliente"])
            if cli_id is None:
                # Tentar inserir cliente novo (da cotação)
                cur.execute(
                    "INSERT INTO clientes (nome) VALUES (%s) ON CONFLICT DO NOTHING RETURNING id, nome",
                    (ct["cliente"],)
                )
                row = cur.fetchone()
                if row:
                    cliente_id[row[1]] = row[0]
                    cli_id = row[0]
                else:
                    cot_skip += 1
                    continue

            cot_batch.append((
                ct["data"],
                cli_id,
                ct.get("origem"),
                ct.get("destino"),
                ct.get("km"),
                ct.get("tipo_veiculo"),
                ct.get("valor_motorista"),
                ct.get("valor_empresa"),
                ct.get("valor_nf"),
                ct["icms_percent"],
                ct["co_percent"],
                ct["imposto_percent"],
                ct["seguro_percent"],
                ct["dias_pagamento"],
                ct["percent_comissao"],
                ct["comissao_valor"],
                ct["lucro"],
                ct["percent_rentabilidade"],
                ct["situacao"],
            ))

        if cot_batch:
            execute_values(cur, """
                INSERT INTO cotacoes (
                    data, cliente_id, origem, destino, km, tipo_veiculo,
                    valor_motorista, valor_empresa, valor_nf,
                    icms_percent, co_percent, imposto_percent, seguro_percent,
                    dias_pagamento, percent_comissao,
                    comissao_valor, lucro, percent_rentabilidade, situacao
                ) VALUES %s
            """, cot_batch)
            cot_ok = len(cot_batch)

        print(f"   {cot_ok} cotacoes inseridas, {cot_skip} ignoradas.")

        conn.commit()
        print("\nMigracao concluida com sucesso!")

        # Resumo final
        cur.execute("SELECT COUNT(*) FROM clientes")
        n_cli = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM motoristas")
        n_mot = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM cargas")
        n_car = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM cotacoes")
        n_cot = cur.fetchone()[0]
        print(f"\nResumo final:")
        print(f"   Clientes:   {n_cli}")
        print(f"   Motoristas: {n_mot}")
        print(f"   Cargas:     {n_car}")
        print(f"   Cotacoes:   {n_cot}")

    except Exception as e:
        conn.rollback()
        print(f"\nERRO: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
